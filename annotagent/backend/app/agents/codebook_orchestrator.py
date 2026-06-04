"""Tool-calling orchestrator for CodebookAgent (Drafter role).

Replaces the single-shot LLM call with a real agent loop that can inspect the
uploaded file directly. The agent gets a small toolbox — list sheets, read
ranges, scan columns for unique values, search for keywords — and decides for
itself how to structure the codebook. This avoids the failure modes of the
one-shot approach (label hallucination on sparse dims, missed sub-categories
in deep sheets, generic codebook names, mode field omissions).

Backend: provider-native tool use. OpenAI uses chat.completions with
``tools`` + ``tool_calls``; Anthropic uses messages content blocks with
``tool_use`` / ``tool_result``.
"""
from __future__ import annotations

import io
import json
import logging
import re
from dataclasses import dataclass, field
from typing import Any

from anthropic import AsyncAnthropic
from openai import AsyncOpenAI

from app.engine.codebook_parser import validate_codebook

logger = logging.getLogger(__name__)


MAX_ITERATIONS = 25
MAX_TOOL_OUTPUT_CHARS = 4000          # truncate tool returns so agents stay in budget
TOOL_OUTPUT_TRUNC_NOTE = "\n\n[... truncated, call again with a narrower range ...]"


# ───────────────────────────────────────────────────────────────────────────
# FileExplorer — wraps the uploaded file and exposes tool methods.
# ───────────────────────────────────────────────────────────────────────────

@dataclass
class FileExplorer:
    """Loads the uploaded file once; tools query the loaded representation."""
    filename: str
    file_bytes: bytes | None = None
    pasted_text: str = ""

    # Lazy-loaded views
    _xlsx_wb: Any = None                              # openpyxl.Workbook
    _text: str | None = None
    _kind: str = ""                                   # "xlsx" | "text" | "csv" | "pdf" | "docx"

    def __post_init__(self) -> None:
        ext = (self.filename.rsplit(".", 1)[-1] or "").lower() if self.filename else ""
        if ext in {"xlsx", "xlsm"} and self.file_bytes:
            import openpyxl
            self._xlsx_wb = openpyxl.load_workbook(io.BytesIO(self.file_bytes), data_only=True, read_only=True)
            self._kind = "xlsx"
        elif self.pasted_text:
            self._text = self.pasted_text
            self._kind = "text"
        elif self.file_bytes:
            # Reuse the existing format parsers for PDF/DOCX/CSV/TXT — agent
            # gets the cleaned text and tools fall back to text search.
            from app.engine.format_parsers import parse_file
            import asyncio
            ingest = asyncio.get_event_loop().run_until_complete(
                parse_file(self.file_bytes, self.filename)
            )
            self._text = ingest.clean_text
            self._kind = ext or "text"

    # ─── Tools ────────────────────────────────────────────────────────────

    def list_sheets(self) -> str:
        """Return sheet names + (rows × cols) for an XLSX, else 'not a spreadsheet'."""
        if self._kind != "xlsx":
            return f"File is {self._kind!r}, not a spreadsheet. Use read_text() or search_text()."
        out = []
        for sn in self._xlsx_wb.sheetnames:
            ws = self._xlsx_wb[sn]
            out.append(f"- {sn!r}: {ws.max_row} rows × {ws.max_column} cols")
        return "\n".join(out)

    def read_sheet_range(self, sheet: str, start_row: int = 1, end_row: int = 30, columns: list[str] | None = None) -> str:
        """Read a row range from a sheet as a markdown-ish table.
        ``columns`` filters by header name (case-insensitive substring match).
        Forward-fills empty cells from the previous non-empty value (handles
        annotator-style sheets where a Coding theme spans multiple level rows).
        """
        if self._kind != "xlsx":
            return "Not a spreadsheet."
        if sheet not in self._xlsx_wb.sheetnames:
            return f"Sheet not found. Available: {self._xlsx_wb.sheetnames}"
        ws = self._xlsx_wb[sheet]
        if start_row < 1: start_row = 1
        if end_row < start_row: end_row = start_row + 30

        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < start_row: continue
            if i > end_row: break
            rows.append(row)
        if not rows:
            return f"(empty range {start_row}–{end_row})"

        header = [str(c or "").strip() for c in rows[0]] if start_row == 1 else None
        if header is None:
            # Pull header from row 1 separately so the table is interpretable.
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            header = [str(c or "").strip() for c in first] if first else [f"col{i}" for i in range(len(rows[0]))]

        # Filter columns by substring match
        keep_idx = list(range(len(header)))
        if columns:
            wanted = [w.lower() for w in columns]
            keep_idx = [i for i, h in enumerate(header) if any(w in (h or "").lower() for w in wanted)]
            if not keep_idx: keep_idx = list(range(len(header)))

        out = [_md_row([header[i] for i in keep_idx])]
        out.append(_md_row(["---"] * len(keep_idx)))

        # Forward-fill across rows so continuation cells make sense
        last = [None] * len(header)
        body_rows = rows[1:] if start_row == 1 else rows
        for row in body_rows:
            row = list(row) + [None] * (len(header) - len(row))
            cells = []
            for i in range(len(header)):
                v = row[i]
                if v is None or (isinstance(v, str) and not v.strip()):
                    v = last[i]
                else:
                    last[i] = v
                cells.append("" if v is None else str(v).strip().replace("\n", " ⏎ ")[:120])
            out.append(_md_row([cells[i] for i in keep_idx]))

        body = "\n".join(out)
        return _truncate(body)

    def column_unique_values(self, sheet: str, column: str, max_values: int = 50) -> str:
        """Return distinct (forward-filled) values in a column with counts.
        Useful for: see what labels exist in 'Levels'; check for ' & ' co-occurrence;
        spot empty rows (continuation pattern signal).
        """
        if self._kind != "xlsx":
            return "Not a spreadsheet."
        if sheet not in self._xlsx_wb.sheetnames:
            return f"Sheet not found. Available: {self._xlsx_wb.sheetnames}"
        ws = self._xlsx_wb[sheet]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            return "(empty sheet)"
        header = [str(c or "").strip().lower() for c in first]
        col = column.strip().lower()
        idx = next((i for i, h in enumerate(header) if col in h), None)
        if idx is None:
            return f"Column not found. Headers: {[c for c in first if c]}"

        counts: dict[str, int] = {}
        ampersand_seen = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[idx] if idx < len(row) else None
            if v is None: continue
            s = str(v).strip()
            if not s: continue
            counts[s] = counts.get(s, 0) + 1
            if " & " in s or "," in s:
                ampersand_seen += 1

        if not counts:
            return "(no values)"
        items = sorted(counts.items(), key=lambda kv: -kv[1])[:max_values]
        out = [f"{n}× {v!r}" for v, n in items]
        out.append(f"\nTotal distinct: {len(counts)}; rows with ' & ' or ',' separator: {ampersand_seen}")
        return _truncate("\n".join(out))

    def search_text(self, query: str, max_results: int = 10) -> str:
        """Substring search across all sheets / pasted text. Case-insensitive.
        Returns hit context: ``[sheet:row]  text…``.
        """
        q = query.lower()
        hits = []
        if self._kind == "xlsx":
            for sn in self._xlsx_wb.sheetnames:
                ws = self._xlsx_wb[sn]
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    for v in row:
                        if v is None: continue
                        s = str(v)
                        if q in s.lower():
                            hits.append(f"[{sn}:{i}]  {s.strip()[:200]}")
                            if len(hits) >= max_results: break
                    if len(hits) >= max_results: break
                if len(hits) >= max_results: break
        elif self._text:
            for i, line in enumerate(self._text.splitlines(), start=1):
                if q in line.lower():
                    hits.append(f"[line:{i}]  {line.strip()[:200]}")
                    if len(hits) >= max_results: break
        if not hits:
            return f"No matches for {query!r}"
        return _truncate("\n".join(hits))

    def read_text(self, max_chars: int = 4000) -> str:
        """Return the cleaned text of the file (for non-XLSX inputs)."""
        if self._kind == "xlsx":
            return "Use list_sheets() / read_sheet_range() for spreadsheets."
        if not self._text:
            return "(no text content)"
        return self._text[:max_chars] + (TOOL_OUTPUT_TRUNC_NOTE if len(self._text) > max_chars else "")


# ───────────────────────────────────────────────────────────────────────────
# Tool schemas (OpenAI function-calling format)
# ───────────────────────────────────────────────────────────────────────────

TOOLS: list[dict[str, Any]] = [
    {
        "type": "function",
        "function": {
            "name": "list_sheets",
            "description": "List all sheets in an XLSX file with their row/column counts. Always call this first if the file is a spreadsheet.",
            "parameters": {"type": "object", "properties": {}, "additionalProperties": False},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_sheet_range",
            "description": "Read a row range from one sheet as a table. Forward-fills empty cells from the previous non-empty value (handles annotator-style sheets where a top-level theme spans many level rows).",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string", "description": "Sheet name (exact, from list_sheets)."},
                    "start_row": {"type": "integer", "description": "First row to read, 1-indexed. Default 1.", "default": 1},
                    "end_row": {"type": "integer", "description": "Last row to read, 1-indexed. Default 30.", "default": 30},
                    "columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Filter to columns whose header contains any of these substrings (case-insensitive).",
                    },
                },
                "required": ["sheet"], "additionalProperties": False,
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "column_unique_values",
            "description": "Distinct values in one column with frequency counts; flags presence of ' & ' or ',' (signals multi-label co-occurrence).",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "column": {"type": "string", "description": "Column name (case-insensitive substring match against headers)."},
                    "max_values": {"type": "integer", "default": 50},
                },
                "required": ["sheet", "column"], "additionalProperties": False,
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_text",
            "description": "Substring search (case-insensitive) across all sheets / pasted text. Returns hit lines with sheet:row context.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string"},
                    "max_results": {"type": "integer", "default": 10},
                },
                "required": ["query"], "additionalProperties": False,
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_text",
            "description": "For non-XLSX inputs (PDF/DOCX/TXT/pasted), return the cleaned text up to max_chars.",
            "parameters": {
                "type": "object",
                "properties": {"max_chars": {"type": "integer", "default": 4000}},
                "additionalProperties": False,
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "propose_codebook",
            "description": "Submit the final codebook JSON. The system validates it; if invalid you'll get an error and should call this again with corrections. Call this when you've finished exploring the file and have a complete schema.",
            "parameters": {
                "type": "object",
                "properties": {
                    "codebook_json": {
                        "type": "string",
                        "description": "STRICT JSON string with: name, description, mode (single_label/multi_label/mixed), dimensions [{name, type, instructions, labels:[{name, definition, examples:[]}]}], decomposition_hints {groups, order}.",
                    },
                },
                "required": ["codebook_json"], "additionalProperties": False,
            },
        },
    },
]


# ───────────────────────────────────────────────────────────────────────────
# System prompt
# ───────────────────────────────────────────────────────────────────────────

ORCHESTRATOR_SYSTEM = """You are CodebookAgent's Drafter role: you build a structured annotation
codebook from messy user input. You have tools to inspect the uploaded file
directly. Use them — do not guess.

Your job, in order:

1. ORIENT — call list_sheets() (if XLSX) or read_text() to see what you're dealing with.
2. INSPECT every relevant sheet/section. Read enough rows to see the actual
   structure. For annotator spreadsheets:
     - top-level themes are usually in column A and span many rows
     - level/subcategory labels are in column B
     - definitions in column C
     - watch for "Emergent codes" sub-tables further down with a different header
3. INFER MODE per dimension:
     - single_label when labels are mutually exclusive (High/Low/No, Yes/No)
     - multi_label when labels can co-occur (you see ' & ' in the data, the
       codebook text says co-occur, or the same item appears with multiple labels)
   Use column_unique_values(...) to check for ' & ' / ',' separators.
4. EXTRACT every distinct label per dimension with its definition. Don't
   invent labels to satisfy a "≥2 labels" rule — if a dimension legitimately
   has only one label, mark it ``binary`` and set labels to
   ``[{name: "Yes, X"}, {name: "No, not X"}]``.
5. NAME the codebook based on the file's actual content (e.g. "AI Companion
   Annotation Codebook", "Self-Disclosure Analysis"). Avoid generic names like
   "Extracted Annotation Codebook".
6. PROPOSE_CODEBOOK with strict JSON. The system validates and tells you what
   to fix if anything's off.

Output schema (when calling propose_codebook):
{
  "name": "...",
  "description": "...",
  "mode": "single_label" | "multi_label" | "mixed",
  "dimensions": [
    {
      "name": "...",
      "type": "single_label" | "multi_label" | "binary",
      "instructions": "...",
      "labels": [{"name": "...", "definition": "...", "examples": []}]
    }
  ],
  "decomposition_hints": {
    "groups": [["dim_a", "dim_b"], ["dim_c"]],
    "order": ["Step 1: …", "Step 2: …"]
  },
  "_rationale_per_dim": {"dim_a": "why I chose this mode + any ambiguity"}
}

Rules:
- Use canonical Title Case for label names.
- Each dimension should have 2-10 labels (binary is OK for a 1-label source).
- Definitions: copy verbatim from the file when present; paraphrase only if
  the source is fragmentary.
- decomposition_hints: group dimensions that share scope (e.g. all
  self-disclosure dims in one step, all AI-behavior dims in another).
- Stop calling exploration tools once you have everything you need —
  unnecessary calls are wasteful."""


# ───────────────────────────────────────────────────────────────────────────
# The agent loop
# ───────────────────────────────────────────────────────────────────────────

@dataclass
class OrchestratorResult:
    ok: bool = False
    codebook: dict[str, Any] = field(default_factory=dict)
    iterations: int = 0
    tool_calls: int = 0
    error: str = ""
    transcript: list[dict[str, Any]] = field(default_factory=list)


async def run_orchestrator(
    *,
    explorer: FileExplorer,
    api_key: str,
    provider: str = "openai",
    model: str = "gpt-5.4-mini",
    max_iterations: int = MAX_ITERATIONS,
    initial_hint: str = "",
) -> OrchestratorResult:
    """Run the tool-calling loop until the agent submits a valid codebook."""
    if (provider or "").lower() == "anthropic":
        return await _run_anthropic_orchestrator(
            explorer=explorer,
            api_key=api_key,
            model=model,
            max_iterations=max_iterations,
            initial_hint=initial_hint,
        )
    return await _run_openai_orchestrator(
        explorer=explorer,
        api_key=api_key,
        model=model,
        max_iterations=max_iterations,
        initial_hint=initial_hint,
    )


async def _run_openai_orchestrator(
    *,
    explorer: FileExplorer,
    api_key: str,
    model: str,
    max_iterations: int,
    initial_hint: str,
) -> OrchestratorResult:
    client = AsyncOpenAI(api_key=api_key)
    messages: list[dict[str, Any]] = [
        {"role": "system", "content": ORCHESTRATOR_SYSTEM},
        {"role": "user", "content": (
            f"File: {explorer.filename!r} (kind={explorer._kind!r}).\n"
            f"{initial_hint}\n\n"
            "Start by orienting yourself — list sheets if XLSX, otherwise read the text. "
            "Then build the codebook and call propose_codebook."
        )},
    ]

    n_tool_calls = 0
    final_codebook: dict[str, Any] | None = None
    last_validation_error = ""

    for it in range(max_iterations):
        try:
            resp = await client.chat.completions.create(
                model=model,
                messages=messages,
                tools=TOOLS,
                tool_choice="auto",
                max_completion_tokens=4096,
            )
        except Exception as e:
            return OrchestratorResult(error=f"OpenAI call failed: {e}", iterations=it, tool_calls=n_tool_calls, transcript=messages)

        msg = resp.choices[0].message
        # Append assistant message verbatim (with any tool_calls).
        messages.append({
            "role": "assistant",
            "content": msg.content or "",
            "tool_calls": [
                {
                    "id": tc.id,
                    "type": "function",
                    "function": {"name": tc.function.name, "arguments": tc.function.arguments},
                }
                for tc in (msg.tool_calls or [])
            ] if msg.tool_calls else None,
        })

        if not msg.tool_calls:
            # No tool call and no propose_codebook — agent stopped without submitting.
            return OrchestratorResult(
                error=f"Agent stopped without calling propose_codebook. Last reply: {(msg.content or '')[:300]}",
                iterations=it + 1, tool_calls=n_tool_calls, transcript=messages,
            )

        # Run each tool call, append a tool message per response
        for tc in msg.tool_calls:
            n_tool_calls += 1
            name = tc.function.name
            try:
                args = json.loads(tc.function.arguments or "{}")
            except json.JSONDecodeError:
                args = {}

            if name == "propose_codebook":
                raw = args.get("codebook_json", "")
                parsed = _safe_json(raw)
                if parsed is None:
                    out = "ERROR: codebook_json was not valid JSON. Return strict JSON object."
                else:
                    errs = list(validate_codebook(parsed))
                    errs.extend(_extra_critic(parsed))
                    if errs:
                        last_validation_error = "; ".join(errs[:5])
                        out = "VALIDATION FAILED:\n- " + "\n- ".join(errs) + "\n\nFix and call propose_codebook again."
                    else:
                        final_codebook = parsed
                        out = "OK: codebook accepted."
            else:
                out = _dispatch(explorer, name, args)

            messages.append({
                "role": "tool",
                "tool_call_id": tc.id,
                "content": _truncate(out),
            })

        if final_codebook is not None:
            return OrchestratorResult(
                ok=True, codebook=final_codebook,
                iterations=it + 1, tool_calls=n_tool_calls, transcript=messages,
            )

    return OrchestratorResult(
        error=f"Hit max_iterations={max_iterations} without a valid codebook. Last validation: {last_validation_error or '(none)'}",
        iterations=max_iterations, tool_calls=n_tool_calls, transcript=messages,
    )


async def _run_anthropic_orchestrator(
    *,
    explorer: FileExplorer,
    api_key: str,
    model: str,
    max_iterations: int,
    initial_hint: str,
) -> OrchestratorResult:
    client = AsyncAnthropic(api_key=api_key)
    messages: list[dict[str, Any]] = [
        {"role": "user", "content": (
            f"File: {explorer.filename!r} (kind={explorer._kind!r}).\n"
            f"{initial_hint}\n\n"
            "Start by orienting yourself — list sheets if XLSX, otherwise read the text. "
            "Then build the codebook and call propose_codebook."
        )},
    ]

    n_tool_calls = 0
    final_codebook: dict[str, Any] | None = None
    last_validation_error = ""
    tools = _anthropic_tools()

    for it in range(max_iterations):
        try:
            resp = await client.messages.create(
                model=model,
                system=ORCHESTRATOR_SYSTEM,
                messages=messages,
                tools=tools,
                tool_choice={"type": "auto"},
                max_tokens=4096,
                temperature=0,
            )
        except Exception as e:
            return OrchestratorResult(error=f"Anthropic call failed: {e}", iterations=it, tool_calls=n_tool_calls, transcript=messages)

        assistant_blocks: list[dict[str, Any]] = []
        tool_uses: list[Any] = []
        for block in resp.content:
            btype = getattr(block, "type", "")
            if btype == "text":
                text = getattr(block, "text", "")
                if text:
                    assistant_blocks.append({"type": "text", "text": text})
            elif btype == "tool_use":
                tool_uses.append(block)
                assistant_blocks.append({
                    "type": "tool_use",
                    "id": block.id,
                    "name": block.name,
                    "input": block.input or {},
                })

        messages.append({"role": "assistant", "content": assistant_blocks})

        if not tool_uses:
            last_text = "\n".join(b.get("text", "") for b in assistant_blocks if b.get("type") == "text")
            return OrchestratorResult(
                error=f"Agent stopped without calling propose_codebook. Last reply: {last_text[:300]}",
                iterations=it + 1, tool_calls=n_tool_calls, transcript=messages,
            )

        tool_results: list[dict[str, Any]] = []
        for tool_use in tool_uses:
            n_tool_calls += 1
            name = tool_use.name
            args = tool_use.input or {}

            if name == "propose_codebook":
                raw = args.get("codebook_json", "")
                parsed = _safe_json(raw)
                if parsed is None:
                    out = "ERROR: codebook_json was not valid JSON. Return strict JSON object."
                else:
                    errs = list(validate_codebook(parsed))
                    errs.extend(_extra_critic(parsed))
                    if errs:
                        last_validation_error = "; ".join(errs[:5])
                        out = "VALIDATION FAILED:\n- " + "\n- ".join(errs) + "\n\nFix and call propose_codebook again."
                    else:
                        final_codebook = parsed
                        out = "OK: codebook accepted."
            else:
                out = _dispatch(explorer, name, args)

            tool_results.append({
                "type": "tool_result",
                "tool_use_id": tool_use.id,
                "content": _truncate(out),
            })

        messages.append({"role": "user", "content": tool_results})

        if final_codebook is not None:
            return OrchestratorResult(
                ok=True, codebook=final_codebook,
                iterations=it + 1, tool_calls=n_tool_calls, transcript=messages,
            )

    return OrchestratorResult(
        error=f"Hit max_iterations={max_iterations} without a valid codebook. Last validation: {last_validation_error or '(none)'}",
        iterations=max_iterations, tool_calls=n_tool_calls, transcript=messages,
    )


def _dispatch(explorer: FileExplorer, name: str, args: dict[str, Any]) -> str:
    """Call the named tool on the explorer with sanitized args."""
    try:
        if name == "list_sheets":
            return explorer.list_sheets()
        if name == "read_sheet_range":
            return explorer.read_sheet_range(
                sheet=args.get("sheet", ""),
                start_row=int(args.get("start_row", 1)),
                end_row=int(args.get("end_row", 30)),
                columns=args.get("columns"),
            )
        if name == "column_unique_values":
            return explorer.column_unique_values(
                sheet=args.get("sheet", ""),
                column=args.get("column", ""),
                max_values=int(args.get("max_values", 50)),
            )
        if name == "search_text":
            return explorer.search_text(
                query=args.get("query", ""),
                max_results=int(args.get("max_results", 10)),
            )
        if name == "read_text":
            return explorer.read_text(max_chars=int(args.get("max_chars", 4000)))
        return f"ERROR: unknown tool {name!r}"
    except Exception as e:
        return f"ERROR running {name!r}: {type(e).__name__}: {e}"


# ───────────────────────────────────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────────────────────────────────

def _md_row(cells: list[str]) -> str:
    return "| " + " | ".join(cells) + " |"


def _truncate(s: str) -> str:
    return s if len(s) <= MAX_TOOL_OUTPUT_CHARS else s[:MAX_TOOL_OUTPUT_CHARS] + TOOL_OUTPUT_TRUNC_NOTE


def _anthropic_tools() -> list[dict[str, Any]]:
    """Convert OpenAI function-tool declarations to Anthropic tool schema."""
    converted = []
    for tool in TOOLS:
        fn = tool["function"]
        converted.append({
            "name": fn["name"],
            "description": fn.get("description", ""),
            "input_schema": fn.get("parameters", {"type": "object", "properties": {}}),
        })
    return converted


def _extra_critic(draft: dict[str, Any]) -> list[str]:
    """Return additional validation errors the orchestrator wants to enforce
    on top of ``validate_codebook``. These are agent-compliance checks: things
    the system prompt asked for but the LLM might forget."""
    errs: list[str] = []
    if not draft.get("mode"):
        errs.append('Top-level "mode" field is missing. Set it to "single_label" / "multi_label" / "mixed".')
    name = (draft.get("name") or "").strip().lower()
    if not name or name in {"codebook", "annotation codebook", "untitled", "untitled codebook", "extracted annotation codebook"}:
        errs.append('Codebook "name" is missing or generic. Pick a content-derived name (e.g. "AI Companion Annotation Codebook").')
    for dim in draft.get("dimensions") or []:
        labels = dim.get("labels") or []
        if len(labels) == 1:
            errs.append(
                f"Dimension {dim.get('name')!r} has only 1 label "
                f"({labels[0].get('name', '?')!r}). Convert to type=binary with two labels: "
                f'{{"name": "Yes, X"}} and {{"name": "No, not X"}}.'
            )
        for lbl in labels:
            if not (lbl.get("definition") or "").strip():
                errs.append(f"Label {lbl.get('name')!r} in dim {dim.get('name')!r} has no definition.")
    return errs


def _safe_json(text: str) -> dict[str, Any] | None:
    s = (text or "").strip()
    if not s: return None
    if s.startswith("```"):
        parts = s.split("\n", 1)
        if len(parts) > 1: s = parts[1]
        s = s.rsplit("```", 1)[0].strip()
    try:
        obj = json.loads(s)
        return obj if isinstance(obj, dict) else None
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", s, flags=re.DOTALL)
        if not m: return None
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            return None
